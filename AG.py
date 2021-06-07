# -*- coding: utf-8 -*-
"""
Created on Tue Apr 20 14:24:21 2021

@author: Pierre
"""
#Lancement

#NOUVEAU TEST


from random import choice, choices, randint, sample, uniform, shuffle
from typing import List, Optional, Callable, Tuple
import numpy as np
from math import *
import time

#-------------------------------- DÉFINITION DES ÉLÉMENTS DE BASE -----------------------------------
#Définition des variables globales

A = 1000
Cp = 100 
hr = 0.1 

DemandeFunc = Callable[[float],float]
RhoFunc = Callable[[float],float]

#Définition des fonction utiles

def demande_1(P:float) -> float:
    '''Cette fonction retourne la valeur de la réponse associée au prix P selon la forme de la demande 1 renseignée dans l'article '''
    #Calcul et retour de la demande associée au prix P
    
    return(-0.003539*(P**3)+2.1215*(P**2)-413.3*P+26580)

def demande_2(P:float) -> float:
    '''Cette fonction retourne la valeur de la réponse associée au prix P selon la forme de la demande 2 renseignée dans l'article '''
    #Calcul et retour de la demande associée au prix P
    
    return(-0.002703*(P**3)+1.577*(P**2)-296.8*P+18413)

def demande_3(P:float) -> float:
    '''Cette fonction retourne la valeur de la réponse associée au prix P selon la forme de la demande 3 renseignée dans l'article '''
    #Calcul et retour de la demande associée au prix P
    
    return(-0.0023*(P**3)+1.35*(P**2)-254.5*P+15500)

def rho_niche(B:float) -> float:
    return(30*(1-exp(-B/1195)))

def rho_masse_1(B:float) -> float:
    return(50*(1-exp(-B/8000)))

def rho_masse_2(B:float) -> float:
    return((1/3.5)*(B**0.5))

def profit_annuel(population : float, demande : DemandeFunc, rho : RhoFunc) -> float:
    '''Cette fonction calcule le profit annuel associé aux paramètres renseignés et au type de demande retenu'''
    
    global A,Cp,hr
    profit = []
    
    #Définition des paramètres utiles
    
    for i in range(len(population)):
        P = population[i][0]
        T = population[i][1]
        B = population[i][2]
        
        D = demande(P) 
        rho_result = rho(B)
        
        Q = T*D
        O = A/T
        C = Cp*D
        H = (hr*Cp*Q)/2 
        profit.append((P+rho_result)*D-O-C-H-B)
    
    return profit

######################################## INITIALISATION ########################################

def initialisation (taille : int):  # definition de la fonction d'initialisation
    return [[uniform(170, 270), uniform(2**(-53), 1), uniform(0, 15000)] for i in range(taille)] # initiation de la population

######################################## SELECTION ########################################

def selection (population : float, taille : int, profit : float):
    
    pop_proba = []
    new_pop = []
    
    pop_profit = list(profit)
    fitness_total = sum(pop_profit)
    
    pop_proba = [pop_profit[i]/fitness_total for i in range(taille)]
    
    for i in range(taille):
        tirage = choices(population,weights=pop_proba,cum_weights=None,k=1)
        if (tirage[0] in new_pop) == False:
            new_pop.extend(tirage)
    
    return new_pop

######################################## MUTATION ########################################

def mutation (population : float, taille : int): # definition de la fonction de mutation
    new_pop = []
    alpha = 0.1
    for i in range(taille):
        alea = uniform(0, 1)
        if alea <= alpha: # selection des 10% de la population pour la mutation
            new_pop.append([uniform(170, 270), uniform(2**(-53), 1), uniform(0, 15000)])
            
    return new_pop

######################################## CROSSOVER ########################################

def crossover (population : float, taille : int):  # definition de la fonction crossover

    parents= []  #initialisation des parents
    enfants = []
    
    for i in range(taille):  # remplir le tableau des parents
        if uniform(0, 1) <= 0.2:
            parents.append([population[i][0], population[i][1], population[i][2]])
            
    if (len(parents) % 2) == 1:
        del parents[-1]
    
    for i in range(1,len(parents),2): # remplir le tableau des enfants 
    
        if randint(1, 2) == 1:
            enfants.append([parents[i-1][0], parents[i][1], parents[i][2]])
            enfants.append([parents[i][0], parents[i - 1][1], parents[i - 1][2]])
        else:
            enfants.append([parents[i-1][0], parents[i - 1][1], parents[i][2]])
            enfants.append([parents[i][0], parents[i][1], parents[i - 1][2]])
            
    return enfants

######################################## AG ########################################

def AG(nb_individu : int, nb_evolution : int, demande : DemandeFunc, rho : RhoFunc):
    
    debut = time.process_time()
    
    ''' Initialisation de la population et récupération du meilleur individu'''
    population = initialisation(nb_individu)
    profit = profit_annuel(population, demande, rho)
    
    #Recherche meilleur individu
    index_best = np.argmax(profit)
    best_profit = profit[index_best]
    best_individu = population[index_best]
    iteration_opt = 0
    
    ''' Evolution de la population afin de récupérer notre meilleure individu après nb_evolution '''
    for i in range(nb_evolution):
        #Selection de la population qui va évoluer
        pop_select = selection(population, len(population), profit)
        
        #Crossover et création des nouveaux individus
        pop_cross = crossover(pop_select, len(pop_select))

        #Mutation 
        pop_mut = mutation(pop_select, len(pop_select))
        
        # On met à jour notre population
        new_pop = []
        new_pop = population + pop_cross + pop_mut
        
        profit_new_pop = profit_annuel(new_pop, demande, rho)
        profit_sorted = sorted(profit_new_pop, reverse = True)
        
            # Selection des nouveaux éléments de la population
        population = [new_pop[profit_new_pop.index(profit_sorted[i])] for i in range(nb_individu)]
        shuffle(population)
        
            #Recherche nouveau meilleur individu
        profit = profit_annuel(population, demande, rho)
        index = np.argmax(profit)
        
        if profit_new_pop[index] > best_profit:
            iteration_opt = i + 1
            index_best = index
            best_profit = profit[index_best]
            best_individu = new_pop[index_best]
            temps_find = time.process_time() - debut
    
    fin = time.process_time() - debut
    # solution = [list(best_individu), best_profit, "Solution trouvée à " + str(find_best)]
    
    print("\n")
    print("La solution a été trouvée au bout de : ", temps_find, " secondes")
    print("L'évolution a été réalisée en : ", fin, " secondes")
    print("\n")
    print("Prix choisi : ", round(best_individu[0], 2), " €")
    print("Périodicité choisi : ", round(best_individu[1], 2))
    print("Budget choisi : ", round(best_individu[2], 2), " €")
    print("\n")
    print("Profit maximal : ", round(best_profit, 2), " €")
    
    return best_individu[0], best_individu[1], best_individu[2], best_profit, iteration_opt, fin

resultat = AG(100, 100, demande_2, rho_niche)

############################################ Sensitivity Analysis ###################################
import openpyxl as op

#Les différentes fonctions fonctionnent avec un fichier DATA.xlsx créé préalablement avec les feuilles correspondantes à
# chaque fonction déjà créées
# Permet de récupérer les données de temps d'exécution en faisant varier le nombre d'individus
def sensibility_time_AG_ind(debut : int, fin : int, pas : int, demande : DemandeFunc, rho : RhoFunc):
    result = []
    for i in range(debut, fin, pas):
        result.append([i, AG(i, 100, demande, rho)[5]])
        print("La solution pour un nb d'individu égale à ", i, "a été trouvée")
    
    fichier = op.load_workbook("DATA.xlsx")
    sheet = fichier['Temps execution petit']
    for i in range(len(result)):
        sheet.cell(3 + i, 1).value = result[i][0]
        sheet.cell(3 + i, 2).value = result[i][1]
    
    fichier.save("DATA.xlsx")
    
    return result

# Permet de récupérer les données de temps d'exécution en faisant varier le nombre de générations
def sensibility_time_AG_gen(debut : int, fin : int, pas : int, demande : DemandeFunc, rho : RhoFunc):
    result = []
    for i in range(debut, fin, pas):
        result.append([i, AG(100, i, demande, rho)[5]])
        print("La solution pour un nb de génération égale à ", i, "a été trouvée")
    
    fichier = op.load_workbook("DATA.xlsx")
    sheet = fichier['Temps execution petit']
    for i in range(len(result)):
        sheet.cell(3 + i, 1).value = result[i][0]
        sheet.cell(3 + i, 3).value = result[i][1]
    
    fichier.save("DATA.xlsx")
    
    return result

# Permet d'exécuter n instances de chaque combinaison de scénarios possibles
def gap_ag(nb_individu : int, nb_gen : int, nb_rep : int):
    result = np.array([[0]*54]*nb_rep, float)
    func_dem = [demande_1, demande_2, demande_3]
    func_rho = [rho_niche, rho_masse_1, rho_masse_2]
    
    posit = 0
    
    for i, dem in enumerate(func_dem):
        for j, r in enumerate(func_rho):
            for k in range(nb_rep):
                temp = AG(100, 100, dem, r)
                result[k][0 + posit] = temp[0]
                result[k][1 + posit] = temp[1]
                result[k][2 + posit] = temp[2]
                result[k][3 + posit] = temp[3]
                result[k][4 + posit] = temp[4]
                result[k][5 + posit] = temp[5]
            posit = posit + 6
            
    fichier = op.load_workbook("DATA.xlsx")
    sheet = fichier['GAP - AG']
    for i in range(nb_rep):
        for j in range(54):
            sheet.cell(3 + i, j + 1).value = result[i][j]
    
    fichier.save("DATA.xlsx")
    return

def iter_opt_AG():
    func_dem = [demande_1, demande_2, demande_3]
    position = 0
    
    result = []
    print("ca c'est AG")
    for i, dem in enumerate(func_dem):
        print(dem)
        for f, gen in enumerate([50, 100, 250, 500]):
            print("Génération : ",gen)
            for j in range(50, 550, 50):
                print("Individu : ",j)
                for k in range(10):
                    temp = AG(j, gen, dem, rho_niche)
                    result.append([i + 1, gen, j, temp[3]])
    
    fichier = op.load_workbook("DATA.xlsx")
    sheet = fichier['GAP OPT AG']
    for i in range(len(result)):
        sheet.cell(1 + i, 1).value = result[i][0]
        sheet.cell(1 + i, 2).value = result[i][1]
        sheet.cell(1 + i, 3).value = result[i][2]
        sheet.cell(1 + i, 4).value = result[i][3]
        
    
    fichier.save("DATA.xlsx")
    
    return 0

# h = iter_opt_AG()
# a = sensibility_time_AG_ind(50, 1050, 50, demande_2, rho_niche)
# b = sensibility_time_AG_gen(50, 1050, 50, demande_2, rho_niche)
# c = gap_ag(100, 100, 1000)

        