# -*- coding: utf-8 -*-
"""
Created on Mon May 17 10:02:30 2021

@author: Pierre
"""

#NOUVEAU TEST


from random import choice, choices, randint, sample, uniform, shuffle
from typing import List, Optional, Callable, Tuple
import numpy as np
from math import *
import time

#-------------------------------- DÉFINITION DES ÉLÉMENTS DE BASE -----------------------------------

#Notation des variables 

# α Maximum additional sales price from product promotion ($)
# β Additional sales price and promotional expense elasticity
# γ Scale parameter used in defining the relationship between the additional sales price and product promotional expense
# A Ordering cost per order ($ per order)
# B Promotional budget ($)
# C Annual purchasing cost ($/year)
# H Annual holding cost ($/year)
# O Annual ordering cost ($/year)
# P Realizable selling price per unit before promotion ($/unit)
# Q Lot size (units)
# R Annual revenue ($/year)
# T Cycle time in year (year)
# D(P) Annual demand at selling price P without any promotion (units/year)
# Cp Unit purchasing cost ($)
# hr Holding cost per monetary unit for a year for the product ($/$-year)
# ρ(B) Additional realizable selling price per unit as a function of expense (B) on product promotion ($)
# Tp Annual profit ($)

#Définition des variables globales

A = 1000
Cp = 100 
hr = 0.1 

DemandeFunc = Callable[[float],float]
RhoFunc = Callable[[float],float]

#Définition des fonction utiles

def demande_1(P:float) -> float:
    '''Cette fonction retourne la valeur de la réponse associée au prix P selon la forme de la demande 1 renseignée dans l'article '''
    return(-0.003539*(P**3)+2.1215*(P**2)-413.3*P+26580)

def demande_2(P:float) -> float:
    '''Cette fonction retourne la valeur de la réponse associée au prix P selon la forme de la demande 2 renseignée dans l'article '''
    return(-0.002703*(P**3)+1.577*(P**2)-296.8*P+18413)

def demande_3(P:float) -> float:
    '''Cette fonction retourne la valeur de la réponse associée au prix P selon la forme de la demande 3 renseignée dans l'article '''
    return(-0.0023*(P**3)+1.35*(P**2)-254.5*P+15500)

def rho_niche(B:float) -> float:
    return(30*(1-exp(-B/1195)))

def rho_masse_1(B:float) -> float:
    return(50*(1-exp(-B/8000)))

def rho_masse_2(B:float) -> float:
    return((1/3.5)*(B**0.5))

def profit_annuel(population : float, demande : DemandeFunc, rho : RhoFunc) -> float:
    '''Cette fonction calcule le profit annuel associé aux paramètres renseignés et au type de demande retenu'''
    
    global A,Cp,hr
    profit = []
    
    #Définition des paramètres utiles
    
    for i in range(len(population)):
        P = population[i][0]
        T = population[i][1]
        B = population[i][2]
        
        D = demande(P) 
        rho_func = rho(B)
        
        Q = T*D
        O = A/T
        C = Cp*D
        H = (hr*Cp*Q)/2 
        profit.append((P+rho_func)*D-O-C-H-B)
    
    return profit

def profit_indiv(individu : float, demande : DemandeFunc, rho : RhoFunc):
    global A,Cp,hr
    
    P = individu[0]
    T = individu[1]
    B = individu[2]
    
    D = demande(P) 
    rho_func = rho(B)
    
    Q = T*D
    # if T == 0 :
    #     print("raté", i, T)
    O = A/T
    C = Cp*D
    H = (hr*Cp*Q)/2 
    profit = (P+rho_func)*D-O-C-H-B
    
    return profit
    

def initialisation (taille : int):  # definition de la fonction d'initialisation
    return [[uniform(170, 270), uniform(2**(-53), 1), uniform(2**(-53), 15000)] for i in range(taille)] # initiation de la population

def teacher_phase(pop : float, profit : float, demande : DemandeFunc, rho : RhoFunc):
    index_teacher = np.argmax(profit)
    teacher = pop[index_teacher]
    size = len(pop)
    
    mean_P = np.mean([pop[i][0] for i in range(size)])
    mean_T = np.mean([pop[i][1] for i in range(size)])
    mean_B = np.mean([pop[i][2] for i in range(size)])
    
    rP = uniform(0, 1)
    rT = uniform(0, 1)
    rB = uniform(0, 1)
    
    for i in range(size):
        
        new_P = pop[i][0] + rP*(teacher[0]-choice([1, 2])*mean_P)
        if new_P < 170:
            new_P = 170
        elif new_P > 270:
            new_P = 270
            
        new_T = pop[i][1] + rT*(teacher[1]-choice([1, 2])*mean_T)
        if new_T < 0:
            new_T = 2**(-53)
        elif new_T > 1:
            new_T = 1
            
        new_B = pop[i][2] + rB*(teacher[2]-choice([1, 2])*mean_B)
        if new_B < 0:
            new_B = 0
        elif new_B > 15000:
            new_B = 15000
        
        new_indiv = [new_P, new_T, new_B]
        new_profit_ind = profit_indiv(new_indiv, demande, rho)
        
        if new_profit_ind > profit[i]:
            pop[i] = new_indiv
            profit[i] = new_profit_ind
            
    return pop, profit

def learning_phase(pop : float, profit : float, demande : DemandeFunc, rho : RhoFunc):
    size = len(pop)
    
    for i in range(size):
        ri = uniform(0, 1)
        j = randint(0, size - 1)
        while j == i:
            j = randint(0, size - 1)
        
        if profit[i] < profit[j]:
            new_P = pop[i][0] + ri*(pop[i][0] - pop[j][0])
            if new_P < 170:
                new_P = 170
            elif new_P > 270:
                new_P = 270
            
            new_T = pop[i][1] + ri*(pop[i][1] - pop[j][1])
            if new_T < 0:
                new_T = 2**(-53)
            elif new_T > 1:
                new_T = 1
            
            new_B = pop[i][2] + ri*(pop[i][2] - pop[j][2])
            if new_B < 0:
                new_B = 0
            elif new_B > 15000:
                new_B = 15000
        else:
            new_P = pop[i][0] + ri*(pop[j][0] - pop[i][0])
            if new_P < 170:
                new_P = 170
            elif new_P > 270:
                new_P = 270
            
            new_T = pop[i][1] + ri*(pop[j][1] - pop[i][1])
            if new_T < 0:
                new_T = 2**(-53)
            elif new_T > 1:
                new_T = 1
            
            new_B = pop[i][2] + ri*(pop[j][2] - pop[i][2])
            if new_B < 0:
                new_B = 0
            elif new_B > 15000:
                new_B = 15000
        
        new_indiv = [new_P, new_T, new_B]
        new_profit_ind = profit_indiv(new_indiv, demande, rho)
        
        if new_profit_ind > profit[i]:
            pop[i] = new_indiv
            profit[i] = new_profit_ind
    
    return pop, profit

def TLBO(nb_individu : int, nb_generation : int, demande : DemandeFunc, rho : RhoFunc):
    #Initialisation de la population
    debut = time.process_time()
    
    population = initialisation(nb_individu)
    profit = profit_annuel(population, demande, rho)
    
    #Recherche du premier meilleur individu
    index_best = np.argmax(profit)
    best_individu = population[index_best]
    best_profit = profit[index_best]
    iteration_opt = 0
    
    for i in range(nb_generation):
        # Phase 1 : Teacher Phase
        pop_teached = teacher_phase(population, profit, demande, rho)
        population = pop_teached[0]
        profit = pop_teached[1]
        
        # Phase 2 : Learning Phase
        pop_learned = learning_phase(population, profit, demande, rho)
        population = pop_learned[0]
        profit = pop_learned[1]
        
        #Recherche nouveau meilleur individu
        index = np.argmax(profit)
        
        if profit[index_best] > best_profit:
            iteration_opt = i + 1
            index_best = index
            best_individu = population[index]
            best_profit = profit[index]
            temps_find = time.process_time() - debut
    
    fin = time.process_time() - debut
    
    print("\n")
    print("La solution a été trouvée au bout de : ", temps_find, " secondes")
    print("L'évolution a été réalisée en : ", fin, " secondes")
    print("L'itération de l'optimalité est : ", iteration_opt)
    print("\n")
    print("Prix choisi : ", best_individu[0], " €")
    print("Périodicité choisi : ", best_individu[1])
    print("Budget choisi : ", best_individu[2], " €")
    print("\n")
    print("Profit maximal : ", round(best_profit, 2), " €")
    
    return best_individu[0], best_individu[1], best_individu[2], best_profit, iteration_opt, fin

TEST24 = TLBO(5000, 100, demande_2, rho_niche)

############################################ Sensitivity Analysis ###################################
import openpyxl as op

def sensibility_time_TLBO_ind(debut : int, fin : int, pas : int, demande : DemandeFunc, rho : RhoFunc):
    result = []
    for i in range(debut, fin, pas):
        result.append([i, TLBO(i, 100, demande, rho)[5]])
        print("La solution pour un nb d'individu égale à ", i, "a été trouvée")
    
    fichier = op.load_workbook("DATA.xlsx")
    sheet = fichier['Temps execution petit']
    for i in range(len(result)):
        # sheet.cell(3 + i, 1).value = result[i][0]
        sheet.cell(3 + i, 4).value = result[i][1]
    
    fichier.save("DATA.xlsx")
    
    return result

def sensibility_time_TLBO_gen(debut : int, fin : int, pas : int, demande : DemandeFunc, rho : RhoFunc):
    result = []
    for i in range(debut, fin, pas):
        result.append([i, TLBO(100, i, demande, rho)[5]])
        print("La solution pour un nb de génération égale à ", i, "a été trouvée")
    
    fichier = op.load_workbook("DATA.xlsx")
    sheet = fichier['Temps execution petit']
    for i in range(len(result)):
        sheet.cell(3 + i, 1).value = result[i][0]
        sheet.cell(3 + i, 5).value = result[i][1]
    
    fichier.save("DATA.xlsx")
    
    return result

def gap_tlbo(nb_individu : int, nb_gen : int, nb_rep : int):
    result = np.array([[0]*54]*nb_rep, float)
    func_dem = [demande_1, demande_2, demande_3]
    func_rho = [rho_niche, rho_masse_1, rho_masse_2]
    
    posit = 0
    
    for i, dem in enumerate(func_dem):
        for j, r in enumerate(func_rho):
            for k in range(nb_rep):
                temp = TLBO(100, 100, dem, r)
                result[k][0 + posit] = temp[0]
                result[k][1 + posit] = temp[1]
                result[k][2 + posit] = temp[2]
                result[k][3 + posit] = temp[3]
                result[k][4 + posit] = temp[4]
                result[k][5 + posit] = temp[5]
            posit = posit + 6
            
    fichier = op.load_workbook("DATA.xlsx")
    sheet = fichier['GAP - TLBO']
    for i in range(nb_rep):
        for j in range(54):
            sheet.cell(3 + i, j + 1).value = result[i][j]
    
    fichier.save("DATA.xlsx")
    return

b = sensibility_time_TLBO_gen(50, 1050, 50, demande_2, rho_niche)
# c = gap_tlbo(100, 100, 1000)
a = sensibility_time_TLBO_ind(50, 1050, 50, demande_2, rho_niche)
